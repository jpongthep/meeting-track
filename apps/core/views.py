import os

from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Q, Max
from django.conf import settings

from apps.core.models import Department, Meeting, FiscalObjective, Section, Command, CommitteeMember

import openpyxl
from openpyxl.utils import get_column_letter

thai_months = [
    "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
    "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."
]

thai_weekdays = [
    "จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"
]

def thai_date(date_obj):
    if not date_obj:
        return ''
    day = date_obj.day
    month = thai_months[date_obj.month - 1]  
    year = date_obj.year + 543  
    
    return f"{day} {month} {year%100}"


def meeting_list(request):
    
    departments = Department.objects.all().order_by("section__type", "section__number")
    meetings = Meeting.objects.select_related('department', 'department__section').all().order_by("department__section__type", "department__section__number")
    
    # departments
    if request.GET.get('q'):
        departments = departments.filter(province__contains=request.GET.get('q'))
            
    if request.GET.get('section'):
        departments = departments.filter(section__name__contains=request.GET.get('section'))
        if request.GET.get('section').isnumeric():
            departments = departments.filter(section__number=request.GET.get('section'))
    
    # meetings
    if request.GET.get('q'):
        meetings = meetings.filter(department__province__contains=request.GET.get('q'))        
            
    if request.GET.get('section'):
        meetings = meetings.filter(department__section__name__contains=request.GET.get('section'))
        if request.GET.get('section').isnumeric():
            meetings = meetings.filter(Q(department__section__number=request.GET.get('section')))
            
    if request.GET.get('year'):
        if request.GET.get('year').isnumeric():
            meetings = meetings.filter(fiscal_year=request.GET.get('year'))
    
    meetings_prepare = (meetings.order_by('department__province')
                                .values('department__province')
                                .filter(status="A"))

    max_meeting_number = meetings_prepare.aggregate(max_meeting_number=Max('number'))['max_meeting_number'] or 1
    meeting_summary = []
    for number in range(1, max_meeting_number + 1):
        meeting_summary.append(meetings_prepare.filter(number=number).count()) 
    
    
    
    context = {
        'departments': departments,
        'meetings_summary': meeting_summary,
        'total': sum(meeting_summary),
    }
    return render(request, 'core/meeting_list.html', context)


def export_meetings(request):


    template_path = os.path.join(settings.BASE_DIR, 'apps', 'core','templates', 'core', 'meetings_template.xlsx')

    wb = openpyxl.load_workbook(template_path)
    sheet = wb['meeting']

    meetings = Meeting.objects.select_related('department', 'department__section').all()
    
    if request.GET.get('q'):
        meetings = meetings.filter(Q(department__province__contains=request.GET.get('q')) | Q(department__section__name__contains=request.GET.get('q')))
        if request.GET.get('q').isnumeric():
            meetings = meetings.filter(Q(department__section__number=request.GET.get('q')))
    
    if request.GET.get('section'):
        meetings = meetings.filter(department__section__name__contains=request.GET.get('section'))
        if request.GET.get('section').isnumeric():
            meetings = meetings.filter(Q(department__section__number=request.GET.get('section')))
                
    if request.GET.get('year'):
        if request.GET.get('year').isnumeric():
            meetings = meetings.filter(fiscal_year=request.GET.get('year'))
            
    meetings = meetings.order_by("department__section__type", "department__section__number",  "department_id", "fiscal_year", "number")
            
    row_num = 3
    
    last_section = None
    last_province = None

    for meeting in meetings:
        department = meeting.department
        section_name = department.section.name
        province = department.province
        fiscal_year = meeting.fiscal_year
        
        meeting_plan = FiscalObjective.objects.filter(department=department, fiscal_year=fiscal_year).first()
        meeting_plan = meeting_plan.meeting_number if meeting_plan else 0
                
        row_data = [
            section_name if last_section != section_name else '',
            province if last_province != province else '',
            f"25{fiscal_year}" if meeting.number == 1 else '',
            meeting_plan if last_province != province else '',
            f"{fiscal_year}-{meeting.number}",
            thai_date(meeting.meeting_date),
            "✔" if meeting.form2_file else '-',
            "✔" if meeting.report_file else '-',
            meeting.summary
        ]
        
        last_section = section_name
        last_province = province
        
        for col_num, value in enumerate(row_data, 1):
            sheet[f'{get_column_letter(col_num)}{row_num}'] = value
        
        row_num += 1   

    meetings =  meetings.filter(status = "A")
    
    max_meeting_number = meetings.aggregate(max_meeting_number=Max('number'))['max_meeting_number'] or 1
    meeting_summary = []
    for number in range(1, max_meeting_number + 1):
        meeting_summary.append(meetings.filter(number=number).count()) 
    
    summary_sheet = wb['summary']

    for i, value in enumerate(meeting_summary):
        cell = summary_sheet.cell(row=3 + i, column=2)  
        cell.value = value

    cell = summary_sheet.cell(row= 3 + len(meeting_summary), column=1)  
    cell.value = "รวม"
    cell = summary_sheet.cell(row= 3 + len(meeting_summary), column=2)  
    cell.value = sum(meeting_summary)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=meetings_export_with_template.xlsx'
    wb.save(response)

    return response


def command_list(request):
    sections = Section.objects.all()
    if request.GET.get('section'):
        sections = sections.filter(name__contains=request.GET.get('section'))
        if request.GET.get('section').isnumeric():
            sections = sections.filter(number=request.GET.get('section'))
    if request.GET.get('q'):
        sections = sections.filter(departments__province__contains=request.GET.get('q'))
    
    for section in sections:
        section.commands = Command.objects.filter(department__section=section).order_by('department_id', '-year', 'number')    
    return render(request, 'core/command_list.html', {'sections': sections})


def committee_members_list(request, command_id):
    command = Command.objects.get(id=command_id)
    committee_members = CommitteeMember.objects.filter(command_id=command_id).order_by('id')

    # Render the template with the committee members
    return render(request, 'core/committee_members_list.html', {'command' : command, 'committee_members': committee_members})